"""
MedPaperHunter - 医学文献智能检索系统主应用
============================================

这是 MedPaperHunter 系统的核心Web服务器文件，使用 FastAPI 框架构建。
该文件负责整合所有后端模块，提供 RESTful API 接口供前端调用。

主要功能模块:
    1. 研究问题分析 - 使用AI分析研究问题并分解为PICO/PEO框架
    2. 检索式生成 - 根据分析结果生成各数据库检索式
    3. 文献检索执行 - 执行PubMed等数据库检索
    4. 文献去重 - 智能去除重复文献
    5. AI筛选 - 使用AI判断文献相关性
    6. 结果导出 - 导出为Excel、CSV、TXT等格式

API端点清单:
    POST /api/search/analyze     - AI分析研究问题，分解为PICO/PEO结构
    POST /api/search/build       - 根据PICO概念生成检索式
    POST /api/search/strategy    - 直接从自然语言生成检索式
    POST /api/search/execute     - 在指定数据库执行检索
    POST /api/search/import-file - 从CSV/Excel文件导入文献
    POST /api/process/dedup      - 去除重复文献
    POST /api/process/screen     - AI相关性筛选
    POST /api/export/excel       - 导出文献到Excel
    POST /api/export/strategies  - 导出检索式到TXT文件
    POST /api/export/csv         - 导出文献到CSV
    GET  /api/config/status      - 检查系统配置状态

技术说明:
    - 使用 FastAPI 作为Web框架，支持自动API文档
    - 使用 CORS 中间件允许跨域请求
    - 静态文件托管前端页面
    - 使用 pydantic 进行数据验证
    - 日志记录系统运行状态

配置要求:
    - 需要配置 LLM_API_KEY 环境变量才能使用AI功能
    - 支持自定义 LLM_BASE_URL 和 LLM_MODEL
    - 输出文件默认保存在项目根目录的 output 文件夹
"""

from __future__ import annotations

import csv
import io
import logging
import os
import tempfile
import time
from typing import Any
from pathlib import Path
from collections import defaultdict

from fastapi import FastAPI, File, HTTPException, UploadFile, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse
from pydantic import BaseModel, Field
from starlette.responses import StreamingResponse
from dotenv import load_dotenv

# Load environment variables from .env file
load_dotenv()

from dedup import deduplicate_by_pmid, deduplicate_by_title, llm_screen
from exporter import articles_to_csv, export_strategies_txt, export_to_excel
from llm_medical_search import LLMClient, LLMConfig, LLMedicalSearchBuilder
from medical_search_strategy_builder import MedicalSearchStrategyBuilder
from pubmed_fetcher import fetch_pubmed

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
)
logger = logging.getLogger("medpaperhunter")

# ---------------------------------------------------------------------------
# 集中配置管理类
# ---------------------------------------------------------------------------

class Config:
    """系统配置管理类。
    
    集中管理所有环境变量配置，提供配置验证功能。
    
    Attributes:
        LLM_API_KEY: LLM API密钥
        LLM_BASE_URL: LLM API基础URL
        LLM_MODEL: LLM模型名称
        OUTPUT_DIR: 输出文件目录
        MAX_FILE_SIZE: 上传文件最大大小（字节）
        MAX_BATCH_SIZE: 批量处理最大数量
        NCBI_REQUEST_DELAY: NCBI API请求间隔（秒）
    """
    # LLM配置
    LLM_API_KEY = os.getenv("LLM_API_KEY", "")
    LLM_BASE_URL = os.getenv("LLM_BASE_URL", "https://api.deepseek.com/v1")
    LLM_MODEL = os.getenv("LLM_MODEL", "deepseek-chat")
    
    # 输出配置
    OUTPUT_DIR = Path(os.getenv("OUTPUT_DIR", str(Path(__file__).parent.parent / "output")))
    
    # 文件上传限制
    MAX_FILE_SIZE = int(os.getenv("MAX_FILE_SIZE", 10 * 1024 * 1024))  # 10MB
    
    # 批量处理配置
    MAX_BATCH_SIZE = int(os.getenv("MAX_BATCH_SIZE", 1000))
    
    # NCBI API配置
    NCBI_REQUEST_DELAY = float(os.getenv("NCBI_REQUEST_DELAY", 0.4))
    
    @classmethod
    def validate(cls):
        """验证配置有效性并记录日志。"""
        if not cls.LLM_API_KEY:
            logger.warning("⚠️ LLM_API_KEY 未配置 - AI功能将被禁用")
        else:
            logger.info("✅ LLM API已配置 (模型: %s)", cls.LLM_MODEL)
        
        # 确保输出目录存在
        cls.OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        logger.info("📁 输出目录: %s", cls.OUTPUT_DIR.resolve())
    
    @classmethod
    def is_llm_configured(cls) -> bool:
        """检查LLM是否已配置。"""
        return bool(cls.LLM_API_KEY)

# 验证配置
Config.validate()

# ---------------------------------------------------------------------------
# FastAPI application
# ---------------------------------------------------------------------------

app = FastAPI(
    title="MedPaperHunter",
    description="Medical literature search, deduplication, screening, and export system",
    version="1.0.0",
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ---------------------------------------------------------------------------
# 请求限流中间件
# ---------------------------------------------------------------------------

# 限流存储：{client_ip: {"count": int, "timestamp": float}}
_request_counts: dict[str, dict[str, float]] = defaultdict(lambda: {"count": 0, "timestamp": 0})
_RATE_LIMIT = 100  # 每分钟最大请求数
_RATE_LIMIT_WINDOW = 60  # 时间窗口（秒）


@app.middleware("http")
async def rate_limit_middleware(request: Request, call_next):
    """请求限流中间件。
    
    限制每个客户端IP每分钟最多请求100次，防止API被滥用。
    
    Args:
        request: HTTP请求对象
        call_next: 下一个中间件或处理函数
    
    Returns:
        HTTP响应
    
    Raises:
        HTTPException: 超过限流阈值时返回429错误
    """
    client_ip = request.client.host if request.client else "unknown"
    now = time.time()
    
    # 获取客户端的请求记录
    record = _request_counts[client_ip]
    
    # 如果时间窗口已过期，重置计数器
    if now - record["timestamp"] > _RATE_LIMIT_WINDOW:
        record["count"] = 0
        record["timestamp"] = now
    
    # 检查是否超过限流阈值
    if record["count"] >= _RATE_LIMIT:
        raise HTTPException(
            status_code=429,
            detail=f"Rate limit exceeded. Please try again in {int(_RATE_LIMIT_WINDOW - (now - record['timestamp']))} seconds.",
        )
    
    # 增加请求计数
    record["count"] += 1
    
    # 调用下一个中间件
    response = await call_next(request)
    
    # 添加限流信息到响应头
    response.headers["X-RateLimit-Limit"] = str(_RATE_LIMIT)
    response.headers["X-RateLimit-Remaining"] = str(_RATE_LIMIT - record["count"])
    
    return response

# Mount frontend static files
frontend_path = Path(__file__).parent.parent / "frontend"
app.mount("/static", StaticFiles(directory=str(frontend_path)), name="static")

# Serve index.html at root
@app.get("/")
async def read_root():
    return FileResponse(str(frontend_path / "index.html"))

# ---------------------------------------------------------------------------
# Shared helpers
# ---------------------------------------------------------------------------


def _get_llm_client() -> LLMClient:
    """创建LLM客户端实例。

    Returns:
        LLMClient: 配置好的LLM客户端实例。

    Raises:
        HTTPException: 如果LLM_API_KEY未配置。
    """
    if not Config.is_llm_configured():
        raise HTTPException(
            status_code=503,
            detail="LLM API key is not configured. Set the LLM_API_KEY environment variable.",
        )
    config = LLMConfig(
        api_key=Config.LLM_API_KEY,
        base_url=Config.LLM_BASE_URL,
        model=Config.LLM_MODEL,
    )
    return LLMClient(config)


def _get_llm_builder() -> LLMedicalSearchBuilder:
    """创建医学检索式构建器实例。

    Returns:
        LLMedicalSearchBuilder: 配置好的检索式构建器实例。

    Raises:
        HTTPException: 如果LLM_API_KEY未配置。
    """
    if not Config.is_llm_configured():
        raise HTTPException(
            status_code=503,
            detail="LLM API key is not configured. Set the LLM_API_KEY environment variable.",
        )
    return LLMedicalSearchBuilder(
        llm_api_key=Config.LLM_API_KEY,
        llm_base_url=Config.LLM_BASE_URL,
        llm_model=Config.LLM_MODEL,
    )


# ---------------------------------------------------------------------------
# Request / Response models
# ---------------------------------------------------------------------------


class AnalyzeRequest(BaseModel):
    question: str = Field(..., description="Natural language research question")
    databases: list[str] = Field(
        default=["pubmed", "embase", "cochrane", "wos", "arxiv"],
        description="Target databases for strategy generation",
    )


class AnalyzeResponse(BaseModel):
    framework: str = Field(..., description="PICO or PEO")
    research_type: str = Field(..., description="Research type description")
    concepts: dict[str, dict[str, Any]] = Field(..., description="PICO concepts")
    language: str = Field(..., description="Input language (zh/en)")
    english_question: str | None = Field(None, description="Translated English question")
    validation_tips: list[str] = Field(default_factory=list, description="Validation tips")


class BuildRequest(BaseModel):
    concepts: dict[str, dict[str, Any]] = Field(..., description="PICO concepts")
    framework: str = Field(default="PICO", description="Framework type (PICO/PEO)")
    databases: list[str] = Field(
        default=["pubmed", "embase", "cochrane", "wos", "arxiv"],
        description="Target databases",
    )
    question: str = Field(default="", description="Original research question")


class BuildResponse(BaseModel):
    strategies: dict[str, str] = Field(..., description="Database-specific search strategies")
    validation_tips: list[str] = Field(default_factory=list, description="Validation tips")
    language: str = Field(default="en", description="Input language")


class StrategyRequest(BaseModel):
    question: str = Field(..., description="Natural language research question")
    databases: list[str] = Field(
        default=["pubmed", "embase", "cochrane", "wos"],
        description="Target databases for strategy generation",
    )
    date_range: str = Field(default="", description="Date range filter, e.g. 2020/01/01-2024/12/31")


class ExecuteRequest(BaseModel):
    strategies: dict[str, str] = Field(..., description="Database-specific search strategies")
    databases: list[str] = Field(..., description="Databases to search")
    date_range: str = Field(default="", description="Date range filter")
    max_results: int = Field(default=500, description="Maximum results per database")


class DedupRequest(BaseModel):
    articles: list[dict[str, Any]] = Field(..., description="List of article dicts to deduplicate")


class ScreenRequest(BaseModel):
    articles: list[dict[str, Any]] = Field(..., description="List of article dicts to screen")
    question: str = Field(..., description="Research question for relevance screening")


class ExportArticlesRequest(BaseModel):
    articles: list[dict[str, Any]] = Field(..., description="List of article dicts to export")


class ExportStrategiesRequest(BaseModel):
    strategies: dict[str, str] = Field(..., description="Database-specific search strategies")





# ---------------------------------------------------------------------------
# API Endpoints
# ---------------------------------------------------------------------------





@app.post("/api/search/analyze", response_model=AnalyzeResponse)
async def analyze_question(request: AnalyzeRequest):
    """Analyze a research question using LLM to perform PICO/PEO decomposition.

    Returns structured PICO concepts that can be edited by the user before
    generating search strategies.
    """
    try:
        builder = _get_llm_builder()

        result = await builder.build_from_question(
            question=request.question,
            databases=request.databases,
        )

        return AnalyzeResponse(
            framework=result["analysis"].get("framework", "PICO"),
            research_type=result["analysis"].get("research_type", "interventional/therapeutic"),
            concepts=result["analysis"].get("concepts", {}),
            language=result.get("language", "en"),
            english_question=result.get("english_question"),
            validation_tips=result.get("validation_tips", []),
        )

    except ValueError as e:
        raise HTTPException(status_code=422, detail=str(e))
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("Failed to analyze question")
        raise HTTPException(status_code=500, detail=f"Failed to analyze question: {e}")


@app.post("/api/search/build", response_model=BuildResponse)
async def build_strategies(request: BuildRequest):
    """Build search strategies from PICO concepts.

    This endpoint accepts user-edited PICO concepts and generates
    database-specific search strategies.
    """
    try:
        strategy_builder = MedicalSearchStrategyBuilder()

        result = strategy_builder.build(
            question=request.question or "User-provided concepts",
            concepts=request.concepts,
            framework=request.framework,
            databases=request.databases,
        )

        return BuildResponse(
            strategies=result.strategies,
            validation_tips=result.validation_tips,
            language=result.language,
        )

    except Exception as e:
        logger.exception("Failed to build strategies")
        raise HTTPException(status_code=500, detail=f"Failed to build strategies: {e}")


@app.post("/api/search/strategy")
async def generate_strategy(request: StrategyRequest):
    """Generate search strategies from a natural language research question.

    Uses LLMedicalSearchBuilder to analyze the question via LLM, validate
    MeSH terms, and produce database-specific search strategies.
    """
    try:
        builder = _get_llm_builder()
        result = await builder.build_from_question(
            question=request.question,
            databases=request.databases,
        )
        return {
            "strategies": result["strategies"],
            "analysis": result["analysis"],
            "mesh_validation": result.get("mesh_validation", {}),
            "validation_tips": result.get("validation_tips", []),
            "language": result.get("language", "en"),
        }
    except ValueError as e:
        raise HTTPException(status_code=422, detail=str(e))
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("Failed to generate search strategy")
        raise HTTPException(status_code=500, detail=f"Failed to generate search strategy: {e}")


@app.post("/api/search/execute")
async def execute_search(request: ExecuteRequest):
    """Execute search on specified databases using provided strategies.

    PubMed is searched via NCBI E-utilities API. Other databases require
    direct network access from an institution with subscription.
    """
    all_articles: list[dict[str, Any]] = []
    counts: dict[str, int] = {}
    errors: list[dict[str, str]] = []

    for db in request.databases:
        strategy = request.strategies.get(db, "")
        if not strategy:
            logger.warning("No strategy provided for database: %s, skipping", db)
            continue

        try:
            if db == "pubmed":
                articles = await fetch_pubmed(
                    query=strategy,
                    max_results=request.max_results,
                    date_range=request.date_range,
                )
                all_articles.extend(articles)
                counts["pubmed"] = len(articles)
                logger.info("PubMed search returned %d articles", len(articles))

            elif db == "scopus":
                # Scopus requires institutional access - currently not implemented
                # Users should use CSV import with exported results
                error_msg = (
                    f"Scopus 检索暂不支持直接API访问。请通过Scopus官网执行检索后，"
                    f"将结果以CSV格式导入系统。Scopus官网: https://www.scopus.com"
                )
                logger.warning(error_msg)
                errors.append({
                    "database": db,
                    "error": error_msg,
                    "suggestion": "请在Scopus官网检索后导入CSV文件"
                })
                counts["scopus"] = 0

            elif db == "wos":
                # Web of Science requires institutional access - currently not implemented
                error_msg = (
                    f"Web of Science 检索暂不支持直接API访问。请通过Web of Science官网执行检索后，"
                    f"将结果以CSV格式导入系统。Web of Science官网: https://www.webofscience.com"
                )
                logger.warning(error_msg)
                errors.append({
                    "database": db,
                    "error": error_msg,
                    "suggestion": "请在Web of Science官网检索后导入CSV文件"
                })
                counts["wos"] = 0

            elif db == "cochrane":
                error_msg = (
                    f"Cochrane Library 检索暂不支持直接API访问。请通过Cochrane官网执行检索后，"
                    f"将结果以CSV格式导入系统。Cochrane Library官网: https://www.cochranelibrary.com"
                )
                logger.warning(error_msg)
                errors.append({
                    "database": db,
                    "error": error_msg,
                    "suggestion": "请在Cochrane Library官网检索后导入CSV文件"
                })
                counts["cochrane"] = 0

            elif db == "arxiv":
                error_msg = (
                    f"arXiv 检索暂不支持直接API访问。请通过arXiv官网执行检索后，"
                    f"将结果以CSV格式导入系统。arXiv官网: https://arxiv.org"
                )
                logger.warning(error_msg)
                errors.append({
                    "database": db,
                    "error": error_msg,
                    "suggestion": "请在arXiv官网检索后导入CSV文件"
                })
                counts["arxiv"] = 0

            else:
                error_msg = f"不支持的数据库: {db}"
                logger.warning(error_msg)
                errors.append({
                    "database": db,
                    "error": error_msg,
                    "suggestion": "请使用支持的数据库或导入CSV文件"
                })
                counts[db] = 0

        except Exception as e:
            logger.exception("Search failed for database: %s", db)
            error_msg = f"{db} 检索失败: {str(e)}"
            errors.append({
                "database": db,
                "error": error_msg,
                "suggestion": "请检查网络连接或稍后重试"
            })
            counts[db] = 0

    return {
        "articles": all_articles,
        "counts": counts,
        "total": len(all_articles),
        "errors": errors,
    }


def _parse_single_file(content: bytes, filename: str) -> list[dict[str, Any]]:
    """解析单个CSV或Excel文件，返回文献列表。
    
    Args:
        content: 文件内容字节
        filename: 文件名（用于判断文件类型）
    
    Returns:
        解析出的文献列表
    
    Raises:
        ValueError: 如果文件格式不支持
    """
    articles: list[dict[str, Any]] = []
    is_xlsx = filename.lower().endswith(".xlsx")
    is_xls = filename.lower().endswith(".xls") and not is_xlsx
    is_csv = filename.lower().endswith(".csv")
    
    reader = []
    if is_xlsx:
        import openpyxl
        from io import BytesIO
        wb = openpyxl.load_workbook(BytesIO(content))
        ws = wb.active
        headers = [str(cell.value).strip().lower() if cell.value else "" for cell in ws[1]]
        rows_data = [[cell.value for cell in row] for row in ws.iter_rows(min_row=2)]
        reader = [dict(zip(headers, row)) for row in rows_data if any(row)]
    elif is_xls:
        # Use xlrd for .xls files (old Excel format)
        import xlrd
        from io import BytesIO
        workbook = xlrd.open_workbook(file_contents=content)
        sheet = workbook.sheet_by_index(0)
        headers = [str(sheet.cell_value(0, i)).strip().lower() if sheet.cell_value(0, i) else "" for i in range(sheet.ncols)]
        for row_idx in range(1, sheet.nrows):
            row_data = [sheet.cell_value(row_idx, i) for i in range(sheet.ncols)]
            if any(row_data):
                reader.append(dict(zip(headers, row_data)))
    elif is_csv:
        text = content.decode("utf-8-sig")
        reader = list(csv.DictReader(io.StringIO(text)))
        # 标准化表头为小写
        for row in reader:
            row_lower = {}
            for k, v in list(row.items()):
                new_key = str(k).strip().lower() if k else ""
                row_lower[new_key] = v
                if new_key != k:
                    del row[k]
            row.update(row_lower)
    else:
        raise ValueError(f"Unsupported file type: {filename}")
    
    logger.info(f"Parsing {filename}: found {len(reader)} rows")
    # Debug: log first 5 rows and headers
    if len(reader) > 0:
        logger.info(f"Headers for {filename}: {list(reader[0].keys())}")
        logger.info(f"First row for {filename}: {reader[0]}")
    
    # 检测Embase多行格式 - 如果第一列包含字段名，第二列包含值
    if is_csv and len(reader) > 0:
        first_row_keys = list(reader[0].keys())
        first_row_values = list(reader[0].values())
        
        # 检查是否是Embase格式：列名中包含字段名关键词，且值也是字段名
        is_embase_format = False
        for key in first_row_keys:
            key_lower = key.lower().strip()
            if key_lower in ['title', 'author names', 'publication year', 'date of publication', 'doi', 'publication type']:
                # 检查值中是否也有字段名
                for val in first_row_values:
                    if val and isinstance(val, str):
                        val_lower = val.lower().strip()
                        if val_lower in ['title', 'author names', 'publication year', 'doi', 'publication type', 'y']:
                            is_embase_format = True
                            break
                if is_embase_format:
                    break
        
        if is_embase_format:
            logger.info(f"Detected Embase multi-line format, using specialized parser...")
            return _parse_embase_csv(content, filename)
    
    # 标准化列名映射表，支持多种常见列名格式
    column_mappings = {
        "source": ["source", "database", "from"],
        "pmid": ["pmid", "pmid.", "pubmed id", "pubmedid", "pubmed-id", "pmid (pubmed)"],
        "title": ["title", "article title", "title.", "document title", "title of article"],
        "journal": ["journal", "journal name", "publication", "source title", "publication title", 
                   "journal/book", "journal / book", "journal or book"],
        "pub_date": ["publication date", "pubdate", "date", "year", "publication year", 
                    "date of publication", "published year"],
        "abstract": ["abstract", "summary", "abstract text", "abstracts"],
        "doi": ["doi", "doi.", "digital object identifier", "doi link"],
        "authors": ["authors", "author", "author list", "author names", "author full names"]
    }
    
    for i, row in enumerate(reader):
        # 使用映射表获取值，支持多种列名格式（忽略大小写和空格）
        def get_value(field_keys):
            for key in field_keys:
                if key in row:
                    return row[key]
            return None
        
        article: dict[str, Any] = {
            "source": get_value(column_mappings["source"]) or "import",
            "pmid": get_value(column_mappings["pmid"]) or None,
            "title": str(get_value(column_mappings["title"]) or "").strip(),
            "authors": [],
            "journal": str(get_value(column_mappings["journal"]) or "").strip(),
            "pub_date": str(get_value(column_mappings["pub_date"]) or "").strip(),
            "abstract": str(get_value(column_mappings["abstract"]) or "").strip(),
            "doi": get_value(column_mappings["doi"]) or None,
        }

        authors_raw = str(get_value(column_mappings["authors"]) or "")
        if authors_raw:
            article["authors"] = [
                a.strip() for a in authors_raw.replace(";", ",").split(",") if a.strip()
            ]

        if article["title"]:
            articles.append(article)
            if i < 5:
                logger.info(f"Parsed article {i+1}: {article['title']}")
    
    logger.info(f"Successfully parsed {len(articles)} articles from {filename}")
    
    return articles


def _parse_embase_csv(content: bytes, filename: str) -> list[dict[str, Any]]:
    """解析Embase导出的特殊CSV格式。
    
    Embase可能导出为多行格式，每行包含字段名和值交替出现。
    例如：
    Title
    Article title here
    Author Names
    Author 1, Author 2
    Publication Year
    2024
    ...
    
    Args:
        content: 文件内容字节
        filename: 文件名
    
    Returns:
        解析出的文献列表
    """
    text = content.decode("utf-8-sig")
    lines = text.strip().split('\n')
    
    articles: list[dict[str, Any]] = []
    current_article: dict[str, Any] | None = None
    current_field: str = ""
    
    # Embase字段名到标准字段的映射
    field_mapping = {
        "title": ["title"],
        "author names": ["authors"],
        "authors": ["authors"],
        "publication year": ["pub_date"],
        "year": ["pub_date"],
        "date of publication": ["pub_date"],
        "doi": ["doi"],
        "publication type": [],
    }
    
    def normalize_field_name(field: str) -> str:
        """规范化字段名"""
        field_lower = field.strip().lower()
        for standard_name, variations in field_mapping.items():
            if field_lower in variations or field_lower == standard_name:
                return standard_name
        return field_lower
    
    def is_field_header(value: str) -> bool:
        """判断是否是字段头（而非值）"""
        value_lower = value.strip().lower()
        field_headers = [
            "title", "author", "year", "publication", "date", "doi", 
            "abstract", "type", "journal", "source"
        ]
        return any(header in value_lower for header in field_headers)
    
    current_article = {"source": "embase", "authors": []}
    
    for line in lines:
        line = line.strip()
        if not line:
            continue
        
        # 尝试解析CSV行
        try:
            parts = list(csv.reader([line]))[0]
        except:
            parts = [line]
        
        if not parts or not parts[0]:
            continue
        
        first_part = parts[0].strip()
        second_part = parts[1].strip() if len(parts) > 1 else ""
        
        # 检测是否是字段头
        normalized_field = normalize_field_name(first_part)
        
        if is_field_header(first_part) and second_part:
            # 如果遇到title字段，说明是新文章开始
            if normalized_field == "title":
                # 保存上一篇文章
                if current_article and current_article.get("title"):
                    articles.append(current_article)
                current_article = {"source": "embase", "authors": [], "title": second_part}
            elif current_article is not None:
                # 记录其他字段
                if normalized_field == "author names" or normalized_field == "authors":
                    authors = [a.strip() for a in second_part.replace(";", ",").split(",") if a.strip()]
                    current_article["authors"] = authors
                elif normalized_field == "publication year" or normalized_field == "year":
                    current_article["pub_date"] = second_part
                elif normalized_field == "date of publication":
                    current_article["pub_date"] = second_part
                elif normalized_field == "doi":
                    current_article["doi"] = second_part if second_part else None
    
    # 保存最后一篇文章
    if current_article and current_article.get("title"):
        articles.append(current_article)
    
    logger.info(f"Parsed {len(articles)} articles from Embase format")
    if articles:
        logger.info(f"First few articles: {[a.get('title', 'NO TITLE') for a in articles[:3]]}")
    
    return articles


@app.post("/api/search/import-file")
async def import_file(files: list[UploadFile] = File(...)):
    """从CSV或Excel文件导入文献（支持多文件同时导入）。
    
    支持从Web of Science、Scopus、Embase等数据库导出的文献文件。
    自动合并多个文件的文献，并确保去重字段（PMID、DOI、标题）正确解析。
    
    支持的文件格式:
        - CSV文件 (.csv)
        - Excel文件 (.xlsx, .xls)
    
    支持的列名（不区分大小写）:
        - Title / article title: 文献标题
        - Authors / author: 作者列表
        - Journal / publication: 期刊名称
        - Publication Date / pubdate / date: 发表日期
        - Abstract / summary: 摘要
        - DOI: 数字对象标识符
        - PMID / pubmed id: PubMed标识符
        - Source / database: 数据来源
    
    文件大小限制: 单个文件最大10MB
    
    Args:
        files: 一个或多个CSV/Excel文件
    
    Returns:
        包含合并后的文献列表和统计信息
    """
    if not files:
        raise HTTPException(status_code=400, detail="No files uploaded.")
    
    all_articles: list[dict[str, Any]] = []
    errors: list[str] = []
    
    for file in files:
        filename = file.filename or "unknown"
        
        # 验证文件类型
        if not (filename.lower().endswith((".csv", ".xlsx", ".xls"))):
            errors.append(f"File '{filename}' has unsupported format. Only CSV/Excel allowed.")
            continue
        
        try:
            # 检查文件大小
            content = await file.read()
            if len(content) > Config.MAX_FILE_SIZE:
                errors.append(f"File '{filename}' exceeds size limit ({Config.MAX_FILE_SIZE // 1024 // 1024}MB).")
                continue
            
            # 解析文件
            articles = _parse_single_file(content, filename)
            logger.info(f"Imported {len(articles)} articles from {filename}")
            all_articles.extend(articles)
        
        except Exception as e:
            logger.exception(f"Failed to import file: {filename}")
            errors.append(f"Failed to parse '{filename}': {str(e)}")
    
    # 合并后进行初步去重（按标题）
    all_articles = deduplicate_by_title(deduplicate_by_pmid(all_articles))
    
    return {
        "articles": all_articles,
        "total_count": len(all_articles),
        "files_uploaded": len(files),
        "errors": errors,
    }


@app.post("/api/process/dedup")
async def deduplicate_articles(request: DedupRequest):
    """Deduplicate articles by PMID first, then by normalized title.

    For each deduplication step, the article with the most complete data
    is retained when duplicates are found.
    """
    if not request.articles:
        return {"articles": [], "removed": 0}

    original_count = len(request.articles)

    # Step 1: Deduplicate by PMID
    after_pmid = deduplicate_by_pmid(request.articles)

    # Step 2: Deduplicate remaining by title
    after_title = deduplicate_by_title(after_pmid)

    removed = original_count - len(after_title)

    return {
        "articles": after_title,
        "removed": removed,
        "original_count": original_count,
        "remaining_count": len(after_title),
    }


@app.post("/api/process/screen")
async def screen_articles(request: ScreenRequest):
    """Screen articles for relevance using LLM.

    Sends article titles and abstracts to the LLM in batches,
    asking whether each article is relevant to the research question.
    """
    if not request.articles:
        return {"articles": [], "total": 0}

    try:
        llm_client = _get_llm_client()
    except HTTPException:
        raise

    try:
        # Get all screened articles with screening_result field
        all_screened = []
        batch_size = 20
        
        from dedup import (
            SCREENING_SYSTEM_PROMPT,
            SCREENING_USER_TEMPLATE,
            _format_article_for_screening,
            _parse_screening_response,
            _call_llm,
        )
        
        for offset in range(0, len(request.articles), batch_size):
            batch = request.articles[offset : offset + batch_size]
            
            # Format articles for the prompt
            articles_text = "\n".join(
                _format_article_for_screening(i + 1, article)
                for i, article in enumerate(batch)
            )
            
            user_message = SCREENING_USER_TEMPLATE.format(
                question=request.question,
                count=len(batch),
                articles_text=articles_text,
            )
            
            messages = [
                {"role": "system", "content": SCREENING_SYSTEM_PROMPT},
                {"role": "user", "content": user_message},
            ]
            
            # Call the LLM
            response_text = await _call_llm(llm_client, messages)
            
            # Parse the response
            screened = _parse_screening_response(response_text, batch)
            all_screened.extend(screened)
        
        # Convert to frontend expected format
        articles_with_screening = []
        for article in all_screened:
            article_copy = dict(article)
            # Map screening_result to screening field (frontend expects this)
            article_copy["screening"] = "included" if article.get("screening_result") == "YES" else "excluded"
            articles_with_screening.append(article_copy)

        included_count = sum(1 for a in articles_with_screening if a["screening"] == "included")
        excluded_count = len(articles_with_screening) - included_count
        
        return {
            "articles": articles_with_screening,
            "total": len(articles_with_screening),
            "included_count": included_count,
            "excluded_count": excluded_count,
        }

    except Exception as e:
        logger.exception("LLM screening failed")
        raise HTTPException(status_code=500, detail=f"LLM screening failed: {e}")


@app.post("/api/export/excel")
async def export_excel(request: ExportArticlesRequest):
    """Export articles to a styled Excel file.

    Returns the Excel file as a downloadable binary stream.
    """
    if not request.articles:
        raise HTTPException(status_code=400, detail="No articles to export.")

    try:
        filepath = str(Config.OUTPUT_DIR / "search_results.xlsx")
        export_to_excel(request.articles, filepath)

        def iterfile():
            with open(filepath, "rb") as f:
                yield from f

        return StreamingResponse(
            iterfile(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=search_results.xlsx"},
        )

    except Exception as e:
        logger.exception("Excel export failed")
        raise HTTPException(status_code=500, detail=f"Excel export failed: {e}")


@app.post("/api/export/strategies")
async def export_strategies(request: ExportStrategiesRequest):
    """Export search strategies to a text file.

    Returns the text file as a downloadable stream.
    """
    if not request.strategies:
        raise HTTPException(status_code=400, detail="No strategies to export.")

    try:
        filepath = str(Config.OUTPUT_DIR / "search_strategies.txt")
        export_strategies_txt(request.strategies, filepath)

        def iterfile():
            with open(filepath, "rb") as f:
                yield from f

        return StreamingResponse(
            iterfile(),
            media_type="text/plain; charset=utf-8",
            headers={"Content-Disposition": "attachment; filename=search_strategies.txt"},
        )

    except Exception as e:
        logger.exception("Strategy export failed")
        raise HTTPException(status_code=500, detail=f"Strategy export failed: {e}")


@app.post("/api/export/csv")
async def export_csv(request: ExportArticlesRequest):
    """Export articles to a CSV file.

    Returns the CSV file as a downloadable stream.
    """
    if not request.articles:
        raise HTTPException(status_code=400, detail="No articles to export.")

    try:
        filepath = str(Config.OUTPUT_DIR / "search_results.csv")
        articles_to_csv(request.articles, filepath)

        def iterfile():
            with open(filepath, "rb") as f:
                yield from f

        return StreamingResponse(
            iterfile(),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": "attachment; filename=search_results.csv"},
        )

    except Exception as e:
        logger.exception("CSV export failed")
        raise HTTPException(status_code=500, detail=f"CSV export failed: {e}")


@app.get("/api/config/status")
async def config_status():
    """检查系统配置状态。
    
    返回当前系统的配置信息，用于前端判断哪些功能可用。
    
    注意事项:
        - API密钥等敏感信息不会暴露，仅返回布尔值表示是否已配置
        - 输出目录只返回路径名，不返回完整绝对路径
    
    Returns:
        dict: 包含LLM配置状态、模型名称和输出目录名
    """
    return {
        "llm_configured": Config.is_llm_configured(),
        "llm_model": Config.LLM_MODEL,
        "output_dir": Config.OUTPUT_DIR.name,  # 只返回目录名，隐藏完整路径
        "max_file_size_mb": Config.MAX_FILE_SIZE // 1024 // 1024,
    }


# ---------------------------------------------------------------------------
# Startup event
# ---------------------------------------------------------------------------


@app.on_event("startup")
async def startup():
    """启动时记录配置状态（不暴露敏感信息）。"""
    logger.info("🚀 MedPaperHunter starting up")
    logger.info("✅ LLM configured: %s (model: %s)", Config.is_llm_configured(), Config.LLM_MODEL)
    logger.info("📁 Output directory: %s", Config.OUTPUT_DIR.resolve())
    logger.info("📦 Max file size: %d MB", Config.MAX_FILE_SIZE // 1024 // 1024)


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    import uvicorn

    uvicorn.run(app, host="0.0.0.0", port=8000)
