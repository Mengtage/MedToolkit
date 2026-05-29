"""
模式选择路由
处理综述模式和RCT模式的选择
"""

from fastapi import APIRouter, Form, File, UploadFile, HTTPException
from fastapi.responses import JSONResponse
from typing import Optional
import io

from services.session_manager import manager, SessionMode
from config.settings import settings

router = APIRouter()


def extract_text_from_file(content: bytes, filename: str) -> str:
    """
    从各种文件格式中提取文本内容
    
    Args:
        content: 文件二进制内容
        filename: 文件名（用于判断扩展名）
        
    Returns:
        提取的文本内容
        
    Raises:
        ValueError: 不支持的文件类型或解析失败
    """
    ext = filename.split('.')[-1].lower() if '.' in filename else ''
    
    if ext in ('txt', 'md'):
        # 文本文件：尝试 UTF-8 和 GBK 编码
        try:
            return content.decode("utf-8")
        except UnicodeDecodeError:
            try:
                return content.decode("gbk")
            except UnicodeDecodeError:
                raise ValueError("文件编码不支持，请使用 UTF-8 或 GBK 编码的文本文件")
    
    elif ext == 'docx':
        # Word 文档
        try:
            from docx import Document
            doc = Document(io.BytesIO(content))
            paragraphs = []
            for para in doc.paragraphs:
                if para.text.strip():
                    paragraphs.append(para.text)
            return '\n'.join(paragraphs)
        except Exception as e:
            raise ValueError(f"无法解析 Word 文档: {str(e)}")
    
    elif ext == 'pdf':
        # PDF 文档
        try:
            import fitz  # PyMuPDF
            doc = fitz.open(stream=content, filetype='pdf')
            pages = []
            for page in doc:
                text = page.get_text()
                if text.strip():
                    pages.append(text)
            doc.close()
            return '\n'.join(pages)
        except Exception as e:
            raise ValueError(f"无法解析 PDF 文档: {str(e)}")
    
    elif ext in ('xlsx', 'xls'):
        # Excel 文档
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            texts = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                sheet_texts = []
                for row in ws.iter_rows(values_only=True):
                    row_text = '\t'.join(str(cell) for cell in row if cell is not None)
                    if row_text.strip():
                        sheet_texts.append(row_text)
                if sheet_texts:
                    texts.append(f"--- Sheet: {sheet_name} ---")
                    texts.extend(sheet_texts)
            wb.close()
            return '\n'.join(texts)
        except Exception as e:
            raise ValueError(f"无法解析 Excel 文档: {str(e)}")
    
    else:
        raise ValueError(f"不支持的文件类型: .{ext}")


@router.post("/select")
async def select_mode(
    mode: str = Form(...),
    research_topic: Optional[str] = Form(None),
    paper_title: Optional[str] = Form(None)
):
    """
    选择写作模式并创建会话

    Args:
        mode: 模式类型 ("review" 或 "rct")
        research_topic: 研究主题（综述模式）
        paper_title: 论文标题（RCT模式）

    Returns:
        session_id: 新创建的会话ID
    """
    if mode not in ["review", "rct"]:
        return JSONResponse({
            "success": False,
            "message": "无效的写作模式"
        }, status_code=400)

    session_mode = SessionMode.RCT if mode == "rct" else SessionMode.REVIEW
    session_id = manager.create_session(session_mode)
    
    # 获取创建的会话以获取user_token
    session = manager.get_session(session_id)
    user_token = session.user_token if session else None

    if session_mode == SessionMode.REVIEW and research_topic:
        manager.update_session(session_id, {"research_topic": research_topic})
    elif session_mode == SessionMode.RCT and paper_title:
        manager.update_session(session_id, {"research_topic": paper_title})

    return JSONResponse({
        "success": True,
        "message": f"{'综述' if mode == 'review' else 'RCT'}模式会话已创建",
        "session_id": session_id,
        "user_token": user_token,  # 返回用户令牌用于后续验证
        "mode": mode
    })


@router.get("/info/{session_id}")
async def get_mode_info(session_id: str):
    """获取会话的模式信息"""
    session = manager.get_session(session_id)
    if not session:
        return JSONResponse({
            "success": False,
            "message": "会话不存在"
        }, status_code=404)

    return JSONResponse({
        "success": True,
        "mode": session.mode.value,
        "status": session.status,
        "research_topic": session.research_topic,
        "research_question": session.research_question
    })


@router.post("/upload/protocol")
async def upload_protocol(
    session_id: str = Form(...),
    protocol: UploadFile = File(...)
):
    """
    上传研究方案（RCT模式）

    Args:
        session_id: 会话ID
        protocol: 研究方案文件

    Returns:
        success: 上传是否成功
    """
    session = manager.get_session(session_id)
    if not session:
        return JSONResponse({
            "success": False,
            "message": "会话不存在"
        }, status_code=404)

    if session.mode != SessionMode.RCT:
        return JSONResponse({
            "success": False,
            "message": "此操作仅适用于RCT模式"
        }, status_code=400)

    try:
        content = await protocol.read()
        filename = protocol.filename or 'protocol.txt'
        protocol_text = extract_text_from_file(content, filename)
        manager.update_session(session_id, {"protocol_text": protocol_text})

        return JSONResponse({
            "success": True,
            "message": "研究方案上传成功",
            "file_size": len(content)
        })
    except ValueError as e:
        return JSONResponse({
            "success": False,
            "message": str(e)
        }, status_code=400)
    except Exception as e:
        return JSONResponse({
            "success": False,
            "message": f"上传失败: {str(e)}"
        }, status_code=500)


@router.post("/upload/analysis")
async def upload_analysis(
    session_id: str = Form(...),
    analysis: UploadFile = File(...)
):
    """
    上传统计分析报告（RCT模式）

    Args:
        session_id: 会话ID
        analysis: 统计分析报告文件

    Returns:
        success: 上传是否成功
    """
    session = manager.get_session(session_id)
    if not session:
        return JSONResponse({
            "success": False,
            "message": "会话不存在"
        }, status_code=404)

    if session.mode != SessionMode.RCT:
        return JSONResponse({
            "success": False,
            "message": "此操作仅适用于RCT模式"
        }, status_code=400)

    try:
        content = await analysis.read()
        filename = analysis.filename or 'analysis.txt'
        analysis_text = extract_text_from_file(content, filename)
        manager.update_session(session_id, {"analysis_text": analysis_text})

        return JSONResponse({
            "success": True,
            "message": "统计分析报告上传成功",
            "file_size": len(content)
        })
    except ValueError as e:
        return JSONResponse({
            "success": False,
            "message": str(e)
        }, status_code=400)
    except Exception as e:
        return JSONResponse({
            "success": False,
            "message": f"上传失败: {str(e)}"
        }, status_code=500)


@router.post("/upload/references")
async def upload_references(
    session_id: str = Form(...),
    references: UploadFile = File(...)
):
    """
    上传参考资料（综述模式）

    Args:
        session_id: 会话ID
        references: 参考资料文件

    Returns:
        上传结果
    """
    session = manager.get_session(session_id)
    if not session:
        return JSONResponse({
            "success": False,
            "message": "会话不存在"
        }, status_code=404)

    try:
        # 安全限制：文件大小最大 10MB
        MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB
        ALLOWED_EXTENSIONS = {'txt', 'md', 'docx', 'pdf', 'xlsx', 'xls'}
        
        # 检查文件扩展名
        filename = references.filename or 'unknown.txt'
        ext = filename.split('.')[-1].lower() if '.' in filename else ''
        if ext not in ALLOWED_EXTENSIONS:
            return JSONResponse({
                "success": False,
                "message": f"不支持的文件类型 .{ext}，仅支持 .txt, .md, .docx, .pdf, .xlsx 文件"
            }, status_code=400)
        
        # 安全文件名处理：移除路径遍历字符
        safe_filename = filename.replace('/', '_').replace('\\', '_').replace('..', '')
        
        # 读取文件内容并检查大小
        content = await references.read()
        if len(content) > MAX_FILE_SIZE:
            return JSONResponse({
                "success": False,
                "message": "文件过大，最大支持 10MB"
            }, status_code=400)
        
        # 使用通用文件解析函数提取文本
        text_content = extract_text_from_file(content, filename)
        
        # 存储参考资料（限制数量）
        MAX_REFERENCES = 10
        if len(session.reference_materials) >= MAX_REFERENCES:
            return JSONResponse({
                "success": False,
                "message": f"最多只能上传 {MAX_REFERENCES} 份参考资料"
            }, status_code=400)
        
        session.reference_materials.append({
            "filename": safe_filename,
            "content": text_content,
            "upload_time": session.updated_at
        })

        return JSONResponse({
            "success": True,
            "message": f"参考资料 {safe_filename} 上传成功",
            "file_size": len(content),
            "total_references": len(session.reference_materials)
        })
    except ValueError as e:
        return JSONResponse({
            "success": False,
            "message": str(e)
        }, status_code=400)
    except Exception as e:
        return JSONResponse({
            "success": False,
            "message": f"上传失败: {str(e)}"
        }, status_code=500)


@router.get("/references/list/{session_id}")
async def list_references(session_id: str):
    """
    获取已上传的参考资料列表

    Args:
        session_id: 会话ID

    Returns:
        参考资料列表
    """
    session = manager.get_session(session_id)
    if not session:
        return JSONResponse({
            "success": False,
            "message": "会话不存在"
        }, status_code=404)

    return JSONResponse({
        "success": True,
        "references": [
            {
                "filename": ref["filename"],
                "upload_time": ref["upload_time"],
                "content_preview": ref["content"][:200] + "..." if len(ref["content"]) > 200 else ref["content"]
            }
            for ref in session.reference_materials
        ]
    })


@router.get("/rct/protocol/generate")
async def rct_protocol_reserved():
    """
    RCT研究方案生成接口（预留）

    此功能暂未开放
    """
    return JSONResponse({
        "status": "reserved",
        "message": "此功能暂未开放，敬请期待",
        "estimated_release": "TBD"
    })
