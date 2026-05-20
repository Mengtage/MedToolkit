"""文献导出模块
==============

本模块负责将文献数据和检索式导出为多种格式：

1. **Excel导出** (.xlsx)
   - 格式化的表格，包含样式美化
   - 自动列宽调整
   - 表头冻结和自动筛选
   - 自动匹配期刊影响因子

2. **CSV导出** (.csv)
   - 简单的文本格式，便于数据交换
   - UTF-8编码，支持中文

3. **检索式导出** (.txt)
   - 格式化的检索式文档
   - 每个数据库的检索式单独列出

核心功能:
    - 影响因子匹配：从CSV文件加载影响因子数据
    - 自动列宽计算：根据内容自动调整列宽
    - 数据格式化：处理列表、None值等特殊情况

导出字段:
    - PMID: PubMed标识符
    - Title: 文献标题
    - Authors: 作者列表
    - Journal: 期刊名称
    - Impact Factor: 影响因子
    - Publication Date: 发表日期
    - Abstract: 摘要
    - DOI: DOI标识符
    - Source: 数据来源
    - Screening Result: 筛选结果
    - Screening Reason: 筛选理由

使用示例:
    >>> # 导出到Excel
    >>> export_to_excel(articles, "results.xlsx")
    >>> 
    >>> # 导出到CSV
    >>> articles_to_csv(articles, "results.csv")
    >>> 
    >>> # 导出检索式
    >>> export_strategies_txt(strategies, "strategies.txt")
"""

from __future__ import annotations

import csv
import logging
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# Header styling constants
HEADER_FILL = PatternFill(start_color="1B2A4A", end_color="1B2A4A", fill_type="solid")
HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
CELL_FONT = Font(name="Arial", size=10)
CELL_ALIGNMENT = Alignment(wrap_text=True, vertical="top")

# Column definitions for Excel export
EXCEL_COLUMNS = [
    ("PMID", "pmid"),
    ("Title", "title"),
    ("Authors", "authors"),
    ("Journal", "journal"),
    ("Impact Factor", "impact_factor"),
    ("Publication Date", "pub_date"),
    ("Abstract", "abstract"),
    ("DOI", "doi"),
    ("Source", "source"),
    ("Screening Result", "screening_result"),
    ("Screening Reason", "screening_reason"),
]

# Load impact factor data
_impact_factors: dict[str, float] = {}


def _load_impact_factors():
    """从CSV文件加载影响因子数据。
    
    影响因子数据文件位于项目根目录的 data/impact_factors.csv
    文件格式: journal,impact_factor
    
    加载策略:
        1. 检查全局缓存是否已加载
        2. 如果未加载，读取CSV文件
        3. 存储到全局字典，键为期刊名小写
    
    Returns:
        影响因子字典，键为期刊名，值为影响因数值
    """
    global _impact_factors
    if _impact_factors:
        return _impact_factors
    
    try:
        # Try to find impact_factors.csv in data directory
        data_dir = Path(__file__).parent.parent / "data"
        csv_path = data_dir / "impact_factors.csv"
        
        if csv_path.exists():
            with open(csv_path, "r", encoding="utf-8") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    journal = row.get("journal", "").strip()
                    if_row = row.get("impact_factor", "").strip()
                    if journal and if_row:
                        try:
                            _impact_factors[journal.lower()] = float(if_row)
                        except ValueError:
                            pass
            logger.info("Loaded %d impact factors", len(_impact_factors))
    except Exception as e:
        logger.warning("Failed to load impact factors: %s", e)
    
    return _impact_factors


def _get_impact_factor(journal: str) -> str:
    """获取期刊的影响因子。
    
    匹配策略:
        1. 精确匹配（期刊名完全一致）
        2. 部分匹配（期刊名包含关系）
    
    Args:
        journal: 期刊名称
    
    Returns:
        影响因子字符串，如果未找到则返回空字符串
    """
    if not journal:
        return ""
    
    # Load if not already loaded
    if not _impact_factors:
        _load_impact_factors()
    
    # Try exact match first
    journal_lower = journal.lower().strip()
    if journal_lower in _impact_factors:
        return str(_impact_factors[journal_lower])
    
    # Try partial match
    for known_journal, if_value in _impact_factors.items():
        if known_journal in journal_lower or journal_lower in known_journal:
            return str(if_value)
    
    return ""


def _format_field(value: Any) -> str:
    """Format an article field value for display.

    Converts lists to comma-separated strings and handles None values.

    Args:
        value: The field value to format.

    Returns:
        Formatted string representation.
    """
    if value is None:
        return ""
    if isinstance(value, list):
        return "; ".join(str(v) for v in value)
    return str(value)


def _calculate_column_widths(
    ws: Any,
    articles: list[dict[str, Any]],
    column_keys: list[str],
) -> dict[str, int]:
    """Calculate appropriate column widths based on content.

    Args:
        ws: The openpyxl worksheet object.
        articles: List of article dicts.
        column_keys: List of dict keys corresponding to columns.

    Returns:
        Dict mapping column keys to recommended widths.
    """
    widths: dict[str, int] = {}

    for col_idx, key in enumerate(column_keys, start=1):
        # Start with header width
        header = ws.cell(row=1, column=col_idx).value or ""
        max_width = len(str(header)) + 2

        # Check content widths (sample first 50 articles for performance)
        for article in articles[:50]:
            value = _format_field(article.get(key, ""))
            # For long text fields, cap the width
            if key in ("abstract", "screening_reason"):
                max_width = max(max_width, min(len(value), 60))
            else:
                max_width = max(max_width, min(len(value), 50))

        widths[key] = max(max_width, 10)

    return widths


def export_to_excel(articles: list[dict[str, Any]], filepath: str) -> None:
    """导出文献到Excel文件。
    
    创建格式化的Excel工作簿，包含样式美化和实用功能。
    
    Excel特性:
        - 表头样式：深蓝色背景 + 白色粗体字
        - 自动列宽调整：根据内容长度自动设置
        - 表头冻结：滚动时表头保持可见
        - 自动筛选：方便数据过滤
        - 自动匹配影响因子
    
    导出列:
        1. PMID - PubMed标识符
        2. Title - 文献标题
        3. Authors - 作者列表（分号分隔）
        4. Journal - 期刊名称
        5. Impact Factor - 影响因子
        6. Publication Date - 发表日期
        7. Abstract - 摘要
        8. DOI - DOI标识符
        9. Source - 数据来源
        10. Screening Result - 筛选结果
        11. Screening Reason - 筛选理由
    
    Args:
        articles: 要导出的文献列表
        filepath: Excel文件输出路径
    """
    logger.info("Exporting %d articles to Excel: %s", len(articles), filepath)

    wb = Workbook()
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet()

    ws.title = "Search Results"

    # Write header row
    column_keys: list[str] = []
    for col_idx, (header_name, key) in enumerate(EXCEL_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        column_keys.append(key)

    # Write data rows
    for row_idx, article in enumerate(articles, start=2):
        for col_idx, key in enumerate(column_keys, start=1):
            # Special handling for impact_factor
            if key == "impact_factor":
                journal = article.get("journal", "")
                value = _get_impact_factor(journal)
            else:
                value = _format_field(article.get(key))
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = CELL_FONT
            cell.alignment = CELL_ALIGNMENT

    # Auto-width columns
    widths = _calculate_column_widths(ws, articles, column_keys)
    for col_idx, key in enumerate(column_keys, start=1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = widths.get(key, 15)

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = ws.dimensions

    wb.save(filepath)
    logger.info("Excel export complete: %s", filepath)


def export_strategies_txt(strategies: dict[str, str], filepath: str) -> None:
    """Write database search strategies to a text file.

    Each strategy is written with a clear header showing the database name,
    followed by the full search query.

    Args:
        strategies: Dict mapping database names to their search query strings.
        filepath: Output file path for the text file (.txt).
    """
    logger.info("Exporting %d search strategies to TXT: %s", len(strategies), filepath)

    separator = "=" * 70

    with open(filepath, "w", encoding="utf-8") as f:
        f.write("MedPaperHunter - Database Search Strategies\n")
        f.write(f"{separator}\n\n")

        for db_name, query in strategies.items():
            f.write(f"Database: {db_name}\n")
            f.write(f"{separator}\n")
            f.write(f"{query}\n")
            f.write("\n\n")

    logger.info("Search strategies export complete: %s", filepath)


def articles_to_csv(articles: list[dict[str, Any]], filepath: str) -> None:
    """Export articles to a CSV file for intermediate storage.

    Uses the same column layout as export_to_excel.

    Args:
        articles: List of article dicts to export.
        filepath: Output file path for the CSV file (.csv).
    """
    logger.info("Exporting %d articles to CSV: %s", len(articles), filepath)

    fieldnames = [key for _, key in EXCEL_COLUMNS]

    with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()

        for article in articles:
            row: dict[str, str] = {}
            for key in fieldnames:
                # Special handling for impact_factor
                if key == "impact_factor":
                    journal = article.get("journal", "")
                    row[key] = _get_impact_factor(journal)
                else:
                    row[key] = _format_field(article.get(key))
            writer.writerow(row)

    logger.info("CSV export complete: %s", filepath)
