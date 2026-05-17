"""
MedPaperHunter - Main FastAPI Application
==========================================

Integrates all backend modules into a complete medical literature search system.

Endpoints:
    POST /api/search/analyze     - LLM analysis + PICO/PEO decomposition (NEW)
    POST /api/search/build       - Generate strategies from PICO concepts (NEW)
    POST /api/search/strategy    - Generate search strategies from natural language
    POST /api/search/execute     - Execute search on specified databases
    POST /api/search/import-csv  - Import articles from CSV
    POST /api/process/dedup      - Deduplicate articles
    POST /api/process/screen     - LLM screening
    POST /api/export/excel        - Export to Excel
    POST /api/export/strategies  - Export strategies to TXT
    POST /api/export/csv         - Export articles to CSV
    GET  /api/config/status      - Check configuration status
"""

from __future__ import annotations

import csv
import io
import logging
import os
import re
import secrets
import tempfile
from typing import Any
from pathlib import Path

from fastapi import FastAPI, File, HTTPException, UploadFile, Request, Header
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse
from pydantic import BaseModel, Field, field_validator
from starlette.responses import StreamingResponse
from dotenv import load_dotenv

# Load environment variables from .env file
load_dotenv()

ALLOWED_ORIGINS = os.getenv("ALLOWED_ORIGINS", "http://localhost:3000,http://localhost:8000,http://localhost:8001").split(",")
MAX_CONTENT_SIZE = int(os.getenv("MAX_CONTENT_SIZE", "10485760"))
RATE_LIMIT_REQUESTS = int(os.getenv("RATE_LIMIT_REQUESTS", "100"))
RATE_LIMIT_WINDOW = int(os.getenv("RATE_LIMIT_WINDOW", "60"))

_request_counts: dict[str, list[float]] = {}

from dedup import deduplicate_by_pmid, deduplicate_by_title, llm_screen
from exporter import articles_to_csv, export_strategies_txt, export_to_excel
from llm_medical_search import LLMClient, LLMConfig, LLMedicalSearchBuilder
from medical_search_strategy_builder import MedicalSearchStrategyBuilder
from pubmed_fetcher import fetch_pubmed

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
)
logger = logging.getLogger("medpaperhunter")

# ---------------------------------------------------------------------------
# Configuration from environment variables
# ---------------------------------------------------------------------------

LLM_API_KEY = os.getenv("LLM_API_KEY", "")
LLM_BASE_URL = os.getenv("LLM_BASE_URL", "https://api.deepseek.com/v1")
LLM_MODEL = os.getenv("LLM_MODEL", "deepseek-chat")
# Set OUTPUT_DIR to project root/output by default
OUTPUT_DIR = os.getenv("OUTPUT_DIR", str(Path(__file__).parent.parent / "output"))

# ---------------------------------------------------------------------------
# Ensure output directory exists
# ---------------------------------------------------------------------------

os.makedirs(OUTPUT_DIR, exist_ok=True)

# ---------------------------------------------------------------------------
# FastAPI application
# ---------------------------------------------------------------------------

app = FastAPI(
    title="MedPaperHunter",
    description="Medical literature search, deduplication, screening, and export system",
    version="1.0.0",
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=ALLOWED_ORIGINS,
    allow_credentials=True,
    allow_methods=["GET", "POST"],
    allow_headers=["*"],
)

# Mount frontend static files
frontend_path = Path(__file__).parent.parent / "frontend"
app.mount("/static", StaticFiles(directory=str(frontend_path)), name="static")

# Serve index.html at root
@app.get("/")
async def read_root():
    return FileResponse(str(frontend_path / "index.html"))

def _check_rate_limit(client_ip: str) -> bool:
    """Check if client has exceeded rate limit."""
    import time
    current_time = time.time()
    if client_ip not in _request_counts:
        _request_counts[client_ip] = []
    _request_counts[client_ip] = [
        t for t in _request_counts[client_ip]
        if current_time - t < RATE_LIMIT_WINDOW
    ]
    if len(_request_counts[client_ip]) >= RATE_LIMIT_REQUESTS:
        return False
    _request_counts[client_ip].append(current_time)
    return True


def sanitize_filename(filename: str) -> str:
    """Sanitize filename to prevent path traversal attacks."""
    filename = re.sub(r'[^\w\s\-\.]', '', filename)
    filename = re.sub(r'\.\.', '', filename)
    return filename[:255]


def _get_client_ip(request: Request) -> str:
    """Extract client IP from request, considering proxy headers."""
    forwarded = request.headers.get("X-Forwarded-For")
    if forwarded:
        return forwarded.split(",")[0].strip()
    real_ip = request.headers.get("X-Real-IP")
    if real_ip:
        return real_ip
    return request.client.host if request.client else "unknown"


def _sanitize_input(value: str, field_name: str, max_length: int = 10000) -> str:
    """Sanitize user input to prevent injection attacks."""
    if not isinstance(value, str):
        return str(value)[:max_length]
    sanitized = value.strip()[:max_length]
    dangerous_patterns = ['<script', 'javascript:', 'onerror=', 'onclick=']
    for pattern in dangerous_patterns:
        sanitized = sanitized.replace(pattern, '')
    return sanitized


# ---------------------------------------------------------------------------
# Shared helpers
# ---------------------------------------------------------------------------


def _get_llm_client() -> LLMClient:
    """Create an LLM client if API key is configured.

    Returns:
        LLMClient instance.

    Raises:
        HTTPException: If LLM_API_KEY is not configured.
    """
    if not LLM_API_KEY:
        raise HTTPException(
            status_code=503,
            detail="LLM API key is not configured. Set the LLM_API_KEY environment variable.",
        )
    config = LLMConfig(
        api_key=LLM_API_KEY,
        base_url=LLM_BASE_URL,
        model=LLM_MODEL,
    )
    return LLMClient(config)


def _get_llm_builder() -> LLMedicalSearchBuilder:
    """Create an LLMedicalSearchBuilder if API key is configured.

    Returns:
        LLMedicalSearchBuilder instance.

    Raises:
        HTTPException: If LLM_API_KEY is not configured.
    """
    if not LLM_API_KEY:
        raise HTTPException(
            status_code=503,
            detail="LLM API key is not configured. Set the LLM_API_KEY environment variable.",
        )
    return LLMedicalSearchBuilder(
        llm_api_key=LLM_API_KEY,
        llm_base_url=LLM_BASE_URL,
        llm_model=LLM_MODEL,
    )


# ---------------------------------------------------------------------------
# Request / Response models
# ---------------------------------------------------------------------------


class AnalyzeRequest(BaseModel):
    question: str = Field(..., description="Natural language research question")
    databases: list[str] = Field(
        default=["pubmed", "embase", "cochrane", "wos", "arxiv"],
        description="Target databases for strategy generation",
    )


class AnalyzeResponse(BaseModel):
    framework: str = Field(..., description="PICO or PEO")
    research_type: str = Field(..., description="Research type description")
    concepts: dict[str, dict[str, Any]] = Field(..., description="PICO concepts")
    language: str = Field(..., description="Input language (zh/en)")
    english_question: str | None = Field(None, description="Translated English question")
    validation_tips: list[str] = Field(default_factory=list, description="Validation tips")


class BuildRequest(BaseModel):
    concepts: dict[str, dict[str, Any]] = Field(..., description="PICO concepts")
    framework: str = Field(default="PICO", description="Framework type (PICO/PEO)")
    databases: list[str] = Field(
        default=["pubmed", "embase", "cochrane", "wos", "arxiv"],
        description="Target databases",
    )
    question: str = Field(default="", description="Original research question")


class BuildResponse(BaseModel):
    strategies: dict[str, str] = Field(..., description="Database-specific search strategies")
    validation_tips: list[str] = Field(default_factory=list, description="Validation tips")
    language: str = Field(default="en", description="Input language")


class StrategyRequest(BaseModel):
    question: str = Field(..., description="Natural language research question")
    databases: list[str] = Field(
        default=["pubmed", "embase", "cochrane", "wos"],
        description="Target databases for strategy generation",
    )
    date_range: str = Field(default="", description="Date range filter, e.g. 2020/01/01-2024/12/31")

    @field_validator('question')
    @classmethod
    def validate_question_length(cls, v: str) -> str:
        if len(v) > 5000:
            raise ValueError('Question must not exceed 5000 characters')
        return v.strip()

    @field_validator('databases')
    @classmethod
    def validate_databases(cls, v: list[str]) -> list[str]:
        allowed = {"pubmed", "embase", "cochrane", "wos", "arxiv", "scopus"}
        return [db for db in v if db in allowed][:10]


class ExecuteRequest(BaseModel):
    strategies: dict[str, str] = Field(..., description="Database-specific search strategies")
    databases: list[str] = Field(..., description="Databases to search")
    date_range: str = Field(default="", description="Date range filter")
    max_results: int = Field(default=500, description="Maximum results per database")

    @field_validator('max_results')
    @classmethod
    def validate_max_results(cls, v: int) -> int:
        if v < 1:
            return 1
        return min(v, 2000)

    @field_validator('databases')
    @classmethod
    def validate_databases(cls, v: list[str]) -> list[str]:
        allowed = {"pubmed", "embase", "cochrane", "wos", "arxiv", "scopus"}
        return [db for db in v if db in allowed][:10]


class DedupRequest(BaseModel):
    articles: list[dict[str, Any]] = Field(..., description="List of article dicts to deduplicate")

    @field_validator('articles')
    @classmethod
    def validate_articles_count(cls, v: list) -> list:
        if len(v) > 10000:
            raise ValueError('Cannot process more than 10000 articles at once')
        return v


class ScreenRequest(BaseModel):
    articles: list[dict[str, Any]] = Field(..., description="List of article dicts to screen")
    question: str = Field(..., description="Research question for relevance screening")

    @field_validator('articles')
    @classmethod
    def validate_articles_count(cls, v: list) -> list:
        if len(v) > 5000:
            raise ValueError('Cannot screen more than 5000 articles at once')
        return v

    @field_validator('question')
    @classmethod
    def validate_question_length(cls, v: str) -> str:
        if len(v) > 5000:
            raise ValueError('Question must not exceed 5000 characters')
        return v.strip()


class ExportArticlesRequest(BaseModel):
    articles: list[dict[str, Any]] = Field(..., description="List of article dicts to export")


class ExportStrategiesRequest(BaseModel):
    strategies: dict[str, str] = Field(..., description="Database-specific search strategies")





# ---------------------------------------------------------------------------
# API Endpoints
# ---------------------------------------------------------------------------





@app.post("/api/search/analyze", response_model=AnalyzeResponse)
async def analyze_question(request: Request, req: AnalyzeRequest):
    """Analyze a research question using LLM to perform PICO/PEO decomposition.

    Returns structured PICO concepts that can be edited by the user before
    generating search strategies.
    """
    client_ip = _get_client_ip(request)
    if not _check_rate_limit(client_ip):
        raise HTTPException(
            status_code=429,
            detail="Too many requests. Please try again later.",
        )

    if not LLM_API_KEY:
        raise HTTPException(
            status_code=503,
            detail="LLM API key is not configured. Set the LLM_API_KEY environment variable.",
        )

    try:
        builder = LLMedicalSearchBuilder(
            llm_api_key=LLM_API_KEY,
            llm_base_url=LLM_BASE_URL,
            llm_model=LLM_MODEL,
        )

        result = await builder.build_from_question(
            question=request.question,
            databases=request.databases,
        )

        return AnalyzeResponse(
            framework=result["analysis"].get("framework", "PICO"),
            research_type=result["analysis"].get("research_type", "interventional/therapeutic"),
            concepts=result["analysis"].get("concepts", {}),
            language=result.get("language", "en"),
            english_question=result.get("english_question"),
            validation_tips=result.get("validation_tips", []),
        )

    except ValueError as e:
        raise HTTPException(status_code=422, detail=str(e))
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("Failed to analyze question")
        raise HTTPException(status_code=500, detail=f"Failed to analyze question: {e}")


@app.post("/api/search/build", response_model=BuildResponse)
async def build_strategies(request: Request, req: BuildRequest):
    """Build search strategies from PICO concepts.

    This endpoint accepts user-edited PICO concepts and generates
    database-specific search strategies.
    """
    client_ip = _get_client_ip(request)
    if not _check_rate_limit(client_ip):
        raise HTTPException(
            status_code=429,
            detail="Too many requests. Please try again later.",
        )

    try:
        strategy_builder = MedicalSearchStrategyBuilder()

        result = strategy_builder.build(
            question=request.question or "User-provided concepts",
            concepts=request.concepts,
            framework=request.framework,
            databases=request.databases,
        )

        return BuildResponse(
            strategies=result.strategies,
            validation_tips=result.validation_tips,
            language=result.language,
        )

    except Exception as e:
        logger.exception("Failed to build strategies")
        raise HTTPException(status_code=500, detail=f"Failed to build strategies: {e}")


@app.post("/api/search/strategy")
async def generate_strategy(request: Request, req: StrategyRequest):
    """Generate search strategies from a natural language research question.

    Uses LLMedicalSearchBuilder to analyze the question via LLM, validate
    MeSH terms, and produce database-specific search strategies.
    """
    client_ip = _get_client_ip(request)
    if not _check_rate_limit(client_ip):
        raise HTTPException(
            status_code=429,
            detail="Too many requests. Please try again later.",
        )

    try:
        builder = _get_llm_builder()
        result = await builder.build_from_question(
            question=req.question,
            databases=req.databases,
        )
        return {
            "strategies": result["strategies"],
            "analysis": result["analysis"],
            "mesh_validation": result.get("mesh_validation", {}),
            "validation_tips": result.get("validation_tips", []),
            "language": result.get("language", "en"),
        }
    except ValueError as e:
        raise HTTPException(status_code=422, detail=str(e))
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("Failed to generate search strategy")
        raise HTTPException(status_code=500, detail=f"Failed to generate search strategy: {e}")


@app.post("/api/search/execute")
async def execute_search(request: Request, req: ExecuteRequest):
    """Execute search on specified databases using provided strategies.

    PubMed is searched via NCBI E-utilities API. Other databases require
    direct network access from an institution with subscription.
    """
    client_ip = _get_client_ip(request)
    if not _check_rate_limit(client_ip):
        raise HTTPException(
            status_code=429,
            detail="Too many requests. Please try again later.",
        )

    all_articles: list[dict[str, Any]] = []
    counts: dict[str, int] = {}
    errors: list[dict[str, str]] = []

    for db in req.databases:
        strategy = req.strategies.get(db, "")
        if not strategy:
            logger.warning("No strategy provided for database: %s, skipping", db)
            continue

        try:
            if db == "pubmed":
                articles = await fetch_pubmed(
                    query=strategy,
                    max_results=req.max_results,
                    date_range=req.date_range,
                )
                all_articles.extend(articles)
                counts["pubmed"] = len(articles)
                logger.info("PubMed search returned %d articles", len(articles))

            elif db == "scopus":
                # Scopus requires institutional access - currently not implemented
                # Users should use CSV import with exported results
                error_msg = (
                    f"Scopus 检索暂不支持直接API访问。请通过Scopus官网执行检索后，"
                    f"将结果以CSV格式导入系统。Scopus官网: https://www.scopus.com"
                )
                logger.warning(error_msg)
                errors.append({
                    "database": db,
                    "error": error_msg,
                    "suggestion": "请在Scopus官网检索后导入CSV文件"
                })
                counts["scopus"] = 0

            elif db == "wos":
                # Web of Science requires institutional access - currently not implemented
                error_msg = (
                    f"Web of Science 检索暂不支持直接API访问。请通过Web of Science官网执行检索后，"
                    f"将结果以CSV格式导入系统。Web of Science官网: https://www.webofscience.com"
                )
                logger.warning(error_msg)
                errors.append({
                    "database": db,
                    "error": error_msg,
                    "suggestion": "请在Web of Science官网检索后导入CSV文件"
                })
                counts["wos"] = 0

            elif db == "cochrane":
                error_msg = (
                    f"Cochrane Library 检索暂不支持直接API访问。请通过Cochrane官网执行检索后，"
                    f"将结果以CSV格式导入系统。Cochrane Library官网: https://www.cochranelibrary.com"
                )
                logger.warning(error_msg)
                errors.append({
                    "database": db,
                    "error": error_msg,
                    "suggestion": "请在Cochrane Library官网检索后导入CSV文件"
                })
                counts["cochrane"] = 0

            elif db == "arxiv":
                error_msg = (
                    f"arXiv 检索暂不支持直接API访问。请通过arXiv官网执行检索后，"
                    f"将结果以CSV格式导入系统。arXiv官网: https://arxiv.org"
                )
                logger.warning(error_msg)
                errors.append({
                    "database": db,
                    "error": error_msg,
                    "suggestion": "请在arXiv官网检索后导入CSV文件"
                })
                counts["arxiv"] = 0

            else:
                error_msg = f"不支持的数据库: {db}"
                logger.warning(error_msg)
                errors.append({
                    "database": db,
                    "error": error_msg,
                    "suggestion": "请使用支持的数据库或导入CSV文件"
                })
                counts[db] = 0

        except Exception as e:
            logger.exception("Search failed for database: %s", db)
            error_msg = f"{db} 检索失败: {str(e)}"
            errors.append({
                "database": db,
                "error": error_msg,
                "suggestion": "请检查网络连接或稍后重试"
            })
            counts[db] = 0

    return {
        "articles": all_articles,
        "counts": counts,
        "total": len(all_articles),
        "errors": errors,
    }


@app.post("/api/search/import-file")
async def import_file(request: Request, file: UploadFile = File(...)):
    """Import articles from a CSV or Excel file.

    Expects columns: Title, Authors, Journal, Publication Date, Abstract, DOI, PMID, Source.
    This endpoint is designed for manually exporting results from other databases.
    """
    client_ip = _get_client_ip(request)
    if not _check_rate_limit(client_ip):
        raise HTTPException(
            status_code=429,
            detail="Too many requests. Please try again later.",
        )

    filename = sanitize_filename(file.filename or "")
    is_excel = filename.lower().endswith((".xlsx", ".xls"))

    if not is_excel and not filename.lower().endswith(".csv"):
        raise HTTPException(status_code=400, detail="Only CSV or Excel files are accepted.")

    try:
        content = await file.read()

        if len(content) > MAX_CONTENT_SIZE:
            raise HTTPException(
                status_code=413,
                detail=f"File size exceeds maximum allowed size of {MAX_CONTENT_SIZE} bytes."
            )

        if is_excel:
            import openpyxl
            from io import BytesIO
            wb = openpyxl.load_workbook(BytesIO(content))
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            rows_data = [[cell.value for cell in row] for row in ws.iter_rows(min_row=2)]
            reader = [dict(zip(headers, row)) for row in rows_data if any(row)]
        else:
            text = content.decode("utf-8-sig")
            reader = list(csv.DictReader(io.StringIO(text)))

        articles: list[dict[str, Any]] = []
        for row in reader:
            row_lower = {k.lower().strip(): v for k, v in row.items()} if row else {}
            article: dict[str, Any] = {
                "source": row.get("Source", "import"),
                "pmid": row.get("PMID") or row.get("pmid") or row.get("PMID".lower()) or None,
                "title": row.get("Title") or row.get("title") or row.get("Title".lower()) or "",
                "authors": [],
                "journal": row.get("Journal", "") or row.get("journal", "") or "",
                "pub_date": row.get("Publication Date", "") or row.get("publication date", "") or "",
                "abstract": row.get("Abstract", "") or row.get("abstract", "") or "",
                "doi": row.get("DOI") or row.get("doi") or row.get("DOI".lower()) or None,
            }

            authors_raw = row.get("Authors", "") or row.get("authors", "") or ""
            if authors_raw:
                article["authors"] = [
                    a.strip() for a in str(authors_raw).replace(";", ",").split(",") if a.strip()
                ]

            if article["title"]:
                articles.append(article)

        return {
            "articles": articles,
            "count": len(articles),
        }

    except Exception as e:
        logger.exception("Failed to import file")
        raise HTTPException(status_code=400, detail=f"Failed to parse file: {e}")


@app.post("/api/process/dedup")
async def deduplicate_articles(request: Request, req: DedupRequest):
    """Deduplicate articles by PMID first, then by normalized title.

    For each deduplication step, the article with the most complete data
    is retained when duplicates are found.
    """
    client_ip = _get_client_ip(request)
    if not _check_rate_limit(client_ip):
        raise HTTPException(
            status_code=429,
            detail="Too many requests. Please try again later.",
        )

    if not req.articles:
        return {"articles": [], "removed": 0}

    original_count = len(req.articles)

    after_pmid = deduplicate_by_pmid(req.articles)
    after_title = deduplicate_by_title(after_pmid)
    removed = original_count - len(after_title)

    return {
        "articles": after_title,
        "removed": removed,
        "original_count": original_count,
        "remaining_count": len(after_title),
    }


@app.post("/api/process/screen")
async def screen_articles(request: Request, req: ScreenRequest):
    """Screen articles for relevance using LLM.

    Sends article titles and abstracts to the LLM in batches,
    asking whether each article is relevant to the research question.
    """
    client_ip = _get_client_ip(request)
    if not _check_rate_limit(client_ip):
        raise HTTPException(
            status_code=429,
            detail="Too many requests. Please try again later.",
        )

    if not req.articles:
        return {"articles": [], "total": 0}

    try:
        llm_client = _get_llm_client()
    except HTTPException:
        raise

    try:
        all_screened = []
        batch_size = 20

        from dedup import (
            SCREENING_SYSTEM_PROMPT,
            SCREENING_USER_TEMPLATE,
            _format_article_for_screening,
            _parse_screening_response,
            _call_llm,
        )

        for offset in range(0, len(req.articles), batch_size):
            batch = req.articles[offset : offset + batch_size]

            articles_text = "\n".join(
                _format_article_for_screening(i + 1, article)
                for i, article in enumerate(batch)
            )

            user_message = SCREENING_USER_TEMPLATE.format(
                question=req.question,
                count=len(batch),
                articles_text=articles_text,
            )
            
            messages = [
                {"role": "system", "content": SCREENING_SYSTEM_PROMPT},
                {"role": "user", "content": user_message},
            ]
            
            # Call the LLM
            response_text = await _call_llm(llm_client, messages)
            
            # Parse the response
            screened = _parse_screening_response(response_text, batch)
            all_screened.extend(screened)
        
        # Convert to frontend expected format
        articles_with_screening = []
        for article in all_screened:
            article_copy = dict(article)
            # Map screening_result to screening field (frontend expects this)
            article_copy["screening"] = "included" if article.get("screening_result") == "YES" else "excluded"
            articles_with_screening.append(article_copy)

        included_count = sum(1 for a in articles_with_screening if a["screening"] == "included")
        excluded_count = len(articles_with_screening) - included_count
        
        return {
            "articles": articles_with_screening,
            "total": len(articles_with_screening),
            "included_count": included_count,
            "excluded_count": excluded_count,
        }

    except Exception as e:
        logger.exception("LLM screening failed")
        raise HTTPException(status_code=500, detail=f"LLM screening failed: {e}")


@app.post("/api/export/excel")
async def export_excel(request: Request, req: ExportArticlesRequest):
    """Export articles to a styled Excel file.

    Returns the Excel file as a downloadable binary stream.
    """
    client_ip = _get_client_ip(request)
    if not _check_rate_limit(client_ip):
        raise HTTPException(
            status_code=429,
            detail="Too many requests. Please try again later.",
        )

    if not req.articles:
        raise HTTPException(status_code=400, detail="No articles to export.")

    try:
        safe_filename = sanitize_filename("search_results.xlsx")
        filepath = os.path.join(OUTPUT_DIR, safe_filename)
        export_to_excel(req.articles, filepath)

        def iterfile():
            with open(filepath, "rb") as f:
                yield from f

        return StreamingResponse(
            iterfile(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={safe_filename}"},
        )

    except Exception as e:
        logger.exception("Excel export failed")
        raise HTTPException(status_code=500, detail="Failed to export Excel file.")


@app.post("/api/export/strategies")
async def export_strategies(request: Request, req: ExportStrategiesRequest):
    """Export search strategies to a text file.

    Returns the text file as a downloadable stream.
    """
    client_ip = _get_client_ip(request)
    if not _check_rate_limit(client_ip):
        raise HTTPException(
            status_code=429,
            detail="Too many requests. Please try again later.",
        )

    if not req.strategies:
        raise HTTPException(status_code=400, detail="No strategies to export.")

    try:
        safe_filename = sanitize_filename("search_strategies.txt")
        filepath = os.path.join(OUTPUT_DIR, safe_filename)
        export_strategies_txt(req.strategies, filepath)

        def iterfile():
            with open(filepath, "rb") as f:
                yield from f

        return StreamingResponse(
            iterfile(),
            media_type="text/plain; charset=utf-8",
            headers={"Content-Disposition": f"attachment; filename={safe_filename}"},
        )

    except Exception as e:
        logger.exception("Strategy export failed")
        raise HTTPException(status_code=500, detail="Failed to export strategies.")


@app.post("/api/export/csv")
async def export_csv(request: Request, req: ExportArticlesRequest):
    """Export articles to a CSV file.

    Returns the CSV file as a downloadable stream.
    """
    client_ip = _get_client_ip(request)
    if not _check_rate_limit(client_ip):
        raise HTTPException(
            status_code=429,
            detail="Too many requests. Please try again later.",
        )

    if not req.articles:
        raise HTTPException(status_code=400, detail="No articles to export.")

    try:
        filepath = os.path.join(OUTPUT_DIR, "search_results.csv")
        articles_to_csv(request.articles, filepath)

        def iterfile():
            with open(filepath, "rb") as f:
                yield from f

        return StreamingResponse(
            iterfile(),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": "attachment; filename=search_results.csv"},
        )

    except Exception as e:
        logger.exception("CSV export failed")
        raise HTTPException(status_code=500, detail=f"CSV export failed: {e}")


@app.get("/api/config/status")
async def config_status():
    """Check what services and features are configured.

    Only returns boolean flags -- API keys and credentials are never exposed.
    """
    return {
        "llm_configured": bool(LLM_API_KEY),
        "llm_model": LLM_MODEL,
        "output_dir": OUTPUT_DIR,
    }


# ---------------------------------------------------------------------------
# Startup event
# ---------------------------------------------------------------------------


@app.on_event("startup")
async def startup():
    """Log configuration status on startup (without exposing secrets)."""
    logger.info("MedPaperHunter starting up")
    logger.info("LLM configured: %s (model: %s)", bool(LLM_API_KEY), LLM_MODEL)
    logger.info("Output directory: %s", os.path.abspath(OUTPUT_DIR))


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    import uvicorn

    uvicorn.run(app, host="0.0.0.0", port=8000)
