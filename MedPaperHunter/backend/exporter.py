"""Exporter module for MedPaperHunter.

Exports articles and search strategies to Excel, CSV, and TXT formats.
"""

from __future__ import annotations

import csv
import logging
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# Header styling constants
HEADER_FILL = PatternFill(start_color="1B2A4A", end_color="1B2A4A", fill_type="solid")
HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
CELL_FONT = Font(name="Arial", size=10)
CELL_ALIGNMENT = Alignment(wrap_text=True, vertical="top")

# Column definitions for Excel export
EXCEL_COLUMNS = [
    ("PMID", "pmid"),
    ("Title", "title"),
    ("Authors", "authors"),
    ("Journal", "journal"),
    ("Impact Factor", "impact_factor"),
    ("Publication Date", "pub_date"),
    ("Abstract", "abstract"),
    ("DOI", "doi"),
    ("Source", "source"),
    ("Screening Result", "screening_result"),
    ("Screening Reason", "screening_reason"),
]

# Load impact factor data
_impact_factors: dict[str, float] = {}


def _load_impact_factors():
    """Load impact factors from CSV file."""
    global _impact_factors
    if _impact_factors:
        return _impact_factors
    
    try:
        # Try to find impact_factors.csv in data directory
        data_dir = Path(__file__).parent.parent / "data"
        csv_path = data_dir / "impact_factors.csv"
        
        if csv_path.exists():
            with open(csv_path, "r", encoding="utf-8") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    journal = row.get("journal", "").strip()
                    if_row = row.get("impact_factor", "").strip()
                    if journal and if_row:
                        try:
                            _impact_factors[journal.lower()] = float(if_row)
                        except ValueError:
                            pass
            logger.info("Loaded %d impact factors", len(_impact_factors))
    except Exception as e:
        logger.warning("Failed to load impact factors: %s", e)
    
    return _impact_factors


def _get_impact_factor(journal: str) -> str:
    """Get impact factor for a journal name."""
    if not journal:
        return ""
    
    # Load if not already loaded
    if not _impact_factors:
        _load_impact_factors()
    
    # Try exact match first
    journal_lower = journal.lower().strip()
    if journal_lower in _impact_factors:
        return str(_impact_factors[journal_lower])
    
    # Try partial match
    for known_journal, if_value in _impact_factors.items():
        if known_journal in journal_lower or journal_lower in known_journal:
            return str(if_value)
    
    return ""


def _format_field(value: Any) -> str:
    """Format an article field value for display.

    Converts lists to comma-separated strings and handles None values.

    Args:
        value: The field value to format.

    Returns:
        Formatted string representation.
    """
    if value is None:
        return ""
    if isinstance(value, list):
        return "; ".join(str(v) for v in value)
    return str(value)


def _calculate_column_widths(
    ws: Any,
    articles: list[dict[str, Any]],
    column_keys: list[str],
) -> dict[str, int]:
    """Calculate appropriate column widths based on content.

    Args:
        ws: The openpyxl worksheet object.
        articles: List of article dicts.
        column_keys: List of dict keys corresponding to columns.

    Returns:
        Dict mapping column keys to recommended widths.
    """
    widths: dict[str, int] = {}

    for col_idx, key in enumerate(column_keys, start=1):
        # Start with header width
        header = ws.cell(row=1, column=col_idx).value or ""
        max_width = len(str(header)) + 2

        # Check content widths (sample first 50 articles for performance)
        for article in articles[:50]:
            value = _format_field(article.get(key, ""))
            # For long text fields, cap the width
            if key in ("abstract", "screening_reason"):
                max_width = max(max_width, min(len(value), 60))
            else:
                max_width = max(max_width, min(len(value), 50))

        widths[key] = max(max_width, 10)

    return widths


def export_to_excel(articles: list[dict[str, Any]], filepath: str) -> None:
    """Export articles to a styled Excel file.

    Creates an Excel workbook with columns: PMID, Title, Authors, Journal,
    Impact Factor, Publication Date, Abstract, DOI, Source, Screening Result, Screening Reason.
    The header row is styled with dark navy background (#1B2A4A) and white text.
    Column widths are auto-adjusted based on content.

    Args:
        articles: List of article dicts to export.
        filepath: Output file path for the Excel file (.xlsx).
    """
    logger.info("Exporting %d articles to Excel: %s", len(articles), filepath)

    wb = Workbook()
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet()

    ws.title = "Search Results"

    # Write header row
    column_keys: list[str] = []
    for col_idx, (header_name, key) in enumerate(EXCEL_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        column_keys.append(key)

    # Write data rows
    for row_idx, article in enumerate(articles, start=2):
        for col_idx, key in enumerate(column_keys, start=1):
            # Special handling for impact_factor
            if key == "impact_factor":
                journal = article.get("journal", "")
                value = _get_impact_factor(journal)
            else:
                value = _format_field(article.get(key))
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = CELL_FONT
            cell.alignment = CELL_ALIGNMENT

    # Auto-width columns
    widths = _calculate_column_widths(ws, articles, column_keys)
    for col_idx, key in enumerate(column_keys, start=1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = widths.get(key, 15)

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = ws.dimensions

    wb.save(filepath)
    logger.info("Excel export complete: %s", filepath)


def export_strategies_txt(strategies: dict[str, str], filepath: str) -> None:
    """Write database search strategies to a text file.

    Each strategy is written with a clear header showing the database name,
    followed by the full search query.

    Args:
        strategies: Dict mapping database names to their search query strings.
        filepath: Output file path for the text file (.txt).
    """
    logger.info("Exporting %d search strategies to TXT: %s", len(strategies), filepath)

    separator = "=" * 70

    with open(filepath, "w", encoding="utf-8") as f:
        f.write("MedPaperHunter - Database Search Strategies\n")
        f.write(f"{separator}\n\n")

        for db_name, query in strategies.items():
            f.write(f"Database: {db_name}\n")
            f.write(f"{separator}\n")
            f.write(f"{query}\n")
            f.write("\n\n")

    logger.info("Search strategies export complete: %s", filepath)


def articles_to_csv(articles: list[dict[str, Any]], filepath: str) -> None:
    """Export articles to a CSV file for intermediate storage.

    Uses the same column layout as export_to_excel.

    Args:
        articles: List of article dicts to export.
        filepath: Output file path for the CSV file (.csv).
    """
    logger.info("Exporting %d articles to CSV: %s", len(articles), filepath)

    fieldnames = [key for _, key in EXCEL_COLUMNS]

    with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()

        for article in articles:
            row: dict[str, str] = {}
            for key in fieldnames:
                # Special handling for impact_factor
                if key == "impact_factor":
                    journal = article.get("journal", "")
                    row[key] = _get_impact_factor(journal)
                else:
                    row[key] = _format_field(article.get(key))
            writer.writerow(row)

    logger.info("CSV export complete: %s", filepath)
